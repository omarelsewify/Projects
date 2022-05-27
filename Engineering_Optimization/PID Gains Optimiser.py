# PID Gains Optimiser 

import numpy as np
import pandas as pd
import openpyxl 
from openpyxl import load_workbook

import matplotlib.pyplot as plt
from numpy.random import permutation as randperm
from scipy.interpolate import Rbf # Import Radial Basis fitting package from scipy
from pytictoc import TicToc

t = TicToc()

plotting = 0

class TrainTest:
    def __init__(self):
        self.reset()
        
    def reset(self):
        self.Train = []
        self.Test = []
    
def train_and_validate(X,y,tt,fit,metric):
    gen_error = metric(fit, X.iloc[tt[1]], y.iloc[tt[1]])
    return gen_error

##################################################### Cross Validation

# K_fold Cross Validation Set Generation
def k_fold_cross_validation_sets(m,k):
    perm = randperm(m)
    kcv_set = TrainTest()
    for i in range(k):
        A = set(range(m))
        B = set(range(i,m,k))
        validate = perm[list(B)]
        train = perm[list(A.difference(B))]
        kcv_set.Train.append(train)
        kcv_set.Test.append(validate)     
    return kcv_set

def cross_validation_estimate(X, y, sets, fit, metric):
    
    tt = [[sets.Train[i], sets.Test[i]] for i in range(len(sets.Test))] 
    error_estimate = [train_and_validate(X,y,tt[i],fit,metric) for i in range(len(tt))]

    return sum(error_estimate)/len(tt)

# Cross Validation Metric
def metric(f, X, y):
    output = sum([(f(X["K_la"].iloc[i], X["x_la"].iloc[i], X["K_long"].iloc[i]) - y.iloc[i])**2 for i in range(len(X)-1)])/len(X)
    return output

##################################################### Import Excel data

# Read in dataset from Excel
workbook = pd.read_excel("Sim3_Results.xlsx", 
            usecols = ['Simulation Mode','K_la', 'x_la', 'K_long',
                       'Max Lateral Error', 'Max Speed Error',
                       'Max Lateral Acceleration','Max Longitudinal Acceleration'],
            engine = 'openpyxl')

# Split data from the 4 Simulation Modes
sim0_data = workbook[workbook['Simulation Mode']==0]    # Nonlinear model
sim1_data = workbook[workbook['Simulation Mode']==1]    # Actuator dynamics only
sim2_data = workbook[workbook['Simulation Mode']==2]    # Actuator dynamics and noise on measurements
sim3_data = workbook[workbook['Simulation Mode']==3]    # Actuator dynamics,noisy measurements, and hold period

# sim_data_all = [sim0_data,sim1_data,sim2_data,sim3_data]
sim_data_all = [sim3_data]

##################################################### Optimising hyperparameter of surrogate model

print('')
sim_number = -1

estimator_functions = []

for sim_data in sim_data_all:
    
    sim_number += 1
    
    # Extract dataset relevant tot each simulation
    X = sim_data[['K_la', "x_la", "K_long"]]
    k_la = sim_data["K_la"]
    x_la = sim_data["x_la"]
    k_long = sim_data["K_long"]

    max_lat_error = sim_data["Max Lateral Error"]
    max_speed_error = sim_data["Max Speed Error"]
    max_ay = sim_data["Max Lateral Acceleration"]
    max_ax = sim_data["Max Longitudinal Acceleration"]

    # Split the dataset into k folds
    sets = TrainTest()
    sets.reset()
    sets = k_fold_cross_validation_sets(len(X),4)   # k = 4

    est_lat_errors = []
    est_speed_errors = []
    est_max_ay = []
    est_max_ax = []

    smooth_params = 10**(np.linspace(start = -10, stop = -2, num=101))

    for smooth_p in smooth_params:
        # Fit Radial Basis Functions to Simulation Mode 0
        #rbf_functions = ["gaussian","inverse","linear"]
        rbfi_lat_error = Rbf(k_la, x_la, k_long, max_lat_error, function = "gaussian", smooth = smooth_p)
        rbfi_speed_error = Rbf(k_la, x_la, k_long, max_speed_error, function = "gaussian", smooth = smooth_p)
        rbfi_max_ay = Rbf(k_la, x_la, k_long, max_ay, function = "gaussian", smooth = smooth_p)
        rbfi_max_ax = Rbf(k_la, x_la, k_long, max_ax, function = "gaussian", smooth = smooth_p)
        
        gen_error_lat_model = cross_validation_estimate(X, max_lat_error, sets, rbfi_lat_error, metric)
        gen_error_speed_model = cross_validation_estimate(X, max_speed_error, sets, rbfi_speed_error, metric)
        gen_max_ay_model = cross_validation_estimate(X, max_ay, sets, rbfi_max_ay, metric)
        gen_max_ax_model = cross_validation_estimate(X, max_ax, sets, rbfi_max_ax, metric)
        
        est_lat_errors.append(gen_error_lat_model)
        est_speed_errors.append(gen_error_speed_model)
        est_max_ay.append(gen_max_ay_model)
        est_max_ax.append(gen_max_ax_model)

    # Optimal Smoothing Parameters
    optimal_smooth_param_lat = smooth_params[np.argmin(est_lat_errors[0:])]
    optimal_smooth_param_speed = smooth_params[np.argmin(est_speed_errors[0:])]
    optimal_smooth_param_max_ay = smooth_params[np.argmin(est_max_ay[0:])]
    optimal_smooth_param_max_ax = smooth_params[np.argmin(est_max_ax[0:])]

    # Optimal functions
    rbfi_lat_error_optimal = Rbf(k_la, x_la, k_long, max_lat_error, function = "gaussian", smooth = optimal_smooth_param_lat)
    rbfi_speed_error_optimal = Rbf(k_la, x_la, k_long, max_speed_error, function = "gaussian", smooth = optimal_smooth_param_speed)
    rbfi_max_ay_optimal =  Rbf(k_la, x_la, k_long, max_ay, function = "gaussian", smooth = optimal_smooth_param_max_ay)
    rbfi_max_ax_optimal =  Rbf(k_la, x_la, k_long, max_ax, function = "gaussian", smooth = optimal_smooth_param_max_ax)

    estimator_functions.append([rbfi_lat_error_optimal,rbfi_speed_error_optimal,rbfi_max_ay_optimal,rbfi_max_ax_optimal])
    
    sim_number = 3
    
    print("Optimal Smoothing Parameter for Lat Error Estimator for Simulator Mode %d = %e " % (sim_number, optimal_smooth_param_lat))
    print("Optimal Smoothing Parameter for Speed Error Estimator for Simulator Mode %d = %e " % (sim_number, optimal_smooth_param_speed))
    print("Optimal Smoothing Parameter for Max a_y Estimator for Simulator Mode %d = %e " % (sim_number, optimal_smooth_param_max_ay))
    print("Optimal Smoothing Parameter for Max a_x Estimator for Simulator Mode %d = %e " % (sim_number, optimal_smooth_param_max_ax))
    print('')
    
    # Plotting of Estimated errors vs Smoothing Parameters
    if plotting:
        plt.title("Simulator Mode %d" % sim_number)
        plt.plot(smooth_params, est_lat_errors, 'g')
        plt.plot(smooth_params, est_speed_errors, 'b')
        plt.legend(["Lateral Error Est.", "Speed Error Est"])
        plt.xlabel("Smoothing Parameter")
        plt.xscale('log')
        plt.ylabel("Estimated Generalisation Error")
        plt.plot(optimal_smooth_param_lat,np.amin(est_lat_errors[0:]),'go')
        plt.plot(optimal_smooth_param_speed,np.amin(est_speed_errors[0:]),'bo')
        plt.show()
    
##################################################### Cross Entropy with Interior Point Method

def cross_entropy_method_interior_point(f, x0, m, m_elite, rho, gamma, tol):
     
    x_history = [x0]

    # Initialising P
    xdata = np.random.rand(m_elite,len(x0))  
    mean = x0
    cov_matrix = np.cov(np.array(xdata).T)
    
    delta = np.Inf
    mean_old = mean
    IPM_sum = []
    
    def c(x):
        output = f(x)        
        
        mle_c = output[0] - 0.2          # Max Lateral error constraint
        mse_c = output[1] - 0.75         # Max Speed error constraint
        may_c = output[2] - 4            # Max Lateral Acceleration
        max_c = output[3] - 4            # Max Longitudinal Acceleration
        
        return [mle_c, mse_c, may_c, max_c]

    # Optimise quadratic penalty function
    while sum(map(lambda i: i > 0, c(mean))) != 0:
        quad_sum = []
        samples = np.random.multivariate_normal(mean, cov_matrix,m)
        
        quad_constraints = [c(x) for x in samples]
        
        for constraint in quad_constraints:
            quad_sum_temp = sum([max(ci,0)**2 for ci in constraint])
            quad_sum.append(quad_sum_temp)
            
        quad_order = np.argsort([x for x in quad_sum])    
        quad_elite_samples = samples[quad_order[0:m_elite]]
        
        mean = np.array([sum(x)/len(x) for x in zip(*quad_elite_samples)])
        
        if np.count_nonzero(c(mean) > 0) == 0:
            x_history.append(mean)
            return x_history
    
    
    while delta > tol:
        
        samples = np.random.multivariate_normal(mean, cov_matrix,m)
        
        IPM_constraints = [c(x) for x in samples]
        
        IPM_sum_temp = 0
        for constraint in IPM_constraints:
            IPM_sum_temp = sum([1/ci for ci in constraint])
            IPM_sum.append(IPM_sum_temp)
        
        min_function = [f(samples[i]) + (1/rho)*IPM_sum[i] for i in range(m)]
   
        IPM_order = np.argsort([(sum(x))for x in min_function])    
        IPM_elite_samples = samples[IPM_order[0:m_elite]]
        
        mean = np.array([sum(x)/len(x) for x in zip(*IPM_elite_samples)])
        
        delta = np.linalg.norm(mean - mean_old)
        mean_old = mean

        rho *= gamma

    return x_history

##################################################### Hooke-Jeeves
def basis(i,n):
    return [int(k == i) for k in range(n)]

def hooke_jeeves(f, x, alpha, eps, gamma, rho):
    
    x_history = []
    n = len(x)
    
    def c(x):
        output = f(x)        
        
        mle_c = max(0.2 - output[0], 0)          # Max Lateral error constraint
        mse_c = max(0.75 - output[1], 0)         # Max Speed error constraint
        may_c = max(4 - output[2], 0)            # Max Lateral Acceleration
        max_c = max(4 - output[3], 0)            # Max Longitudinal Acceleration
        
        # GAINS CANNOT BE NEGATIVE
        K_la_c = max(-x[0], 0)*np.Inf
        x_la_c = max(-x[1], 0)*np.Inf
        K_long_c = max(-x[2], 0)*np.Inf
        
        return [mle_c, mse_c, may_c, max_c, K_la_c, x_la_c, K_long_c]
    
    def obj_function(x,f,c,rho):
        return  sum(f(x) + (sum(map(lambda i: i**2 , c(x))))*rho)
    
    while alpha > eps:
        improved = False
        x_best = x
        obj_best = obj_function(x_best,f,c,rho)
        
        for i in range(n): 
            for sgn in (-1,1):
                vector = [sgn*alpha*idx for idx in basis(i,n)] 
                x_temp = [x[i] + vector[i] for i in range(n)]
                obj_temp = obj_function(x_temp,f,c,rho)
                
                if obj_temp < obj_best:
                    x_best = x_temp
                    x_history.append(x_best)
                    improved = True
        
        x = x_best
        
        if improved == False:
            alpha *= gamma
    
    return x_history

##################################################### Design Point optimisation of the Surrogate Models

sim_number = -1

for sim_mode in estimator_functions:
    
    sim_number += 1

    sim_data = sim_data_all[sim_number]
    # Extract dataset relevant tot each simulation
    X = sim_data[['K_la', "x_la", "K_long"]]
    k_la = sim_data["K_la"]
    x_la = sim_data["x_la"]
    k_long = sim_data["K_long"]

    max_lat_error = sim_data["Max Lateral Error"]
    max_speed_error = sim_data["Max Speed Error"]
    max_ay = sim_data["Max Lateral Acceleration"]
    max_ax = sim_data["Max Longitudinal Acceleration"]
    
    lateral_error_f = sim_mode[0]
    speed_error_f = sim_mode[1]
    max_ay_f = sim_mode[2]
    max_ax_f = sim_mode[3]
    
    def f(x):
        lateral_error = lateral_error_f(x[0],x[1],x[2])
        speed_error = speed_error_f(x[0],x[1],x[2])
        max_ay = max_ay_f(x[0],x[1],x[2])
        max_ax = max_ax_f(x[0],x[1],x[2])

        return [lateral_error.astype(float),speed_error.astype(float), 
                max_ay.astype(float),max_ax.astype(float)]
    
    x0 = [np.random.uniform(min(k_la),max(k_la)),
          np.random.uniform(min(x_la),max(x_la)),
          np.random.uniform(min(k_long),max(k_long))]
    
    # t.tic()
    # x_history = cross_entropy_method_interior_point(f, x0, m = 50, m_elite = 5, rho = 0.1, gamma = 1.2, tol = 0.01)
    # t.toc()
    
    t.tic()
    x_history = hooke_jeeves(f, x0, alpha = 0.5, eps = 0.01 , gamma = 0.5, rho = 5)
    t.toc()
    
    x_best = x_history[-1]
    print("Optimal PID Gains for Simulation Mode %d:" %sim_number)
    print(x_best)
    print("")

    # Save data to excel sheet
    
    data2save = [3, x_best[0], x_best[1], x_best[2]]
    df = pd.DataFrame([data2save])
    
    path = "Sim3_Results.xlsx"
    
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)

    writer.save()
    
    plt.plot(range(len(x_history)),[x[0] for x in x_history])